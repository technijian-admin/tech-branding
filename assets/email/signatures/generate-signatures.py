"""Generate individual email signatures for all Technijian employees."""
import openpyxl
import os
import re

# --- Configuration ---
EXCEL_PATH = r'C:\vscode\tech-branding\tech-branding\data\Technijian Directory  (1).xlsx'
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'employees')
OFFICE_NUMBER = '949.379.8499'
SUPPORT_NUMBER = '949.379.8501'

SEO_BOOKING_URL = 'https://outlook.office.com/book/SEOMeeting@Technijian365.onmicrosoft.com/?ismsaljsauthenabled=true'
DEV_BOOKING_URL = 'https://outlook.office.com/book/Developers@Technijian365.onmicrosoft.com/'
SUPPORT_BOOKING_URL = 'https://outlook.office365.com/owa/calendar/Meetingwithsupport@Technijian365.onmicrosoft.com/bookings/'
LOGO_URL = 'https://technijian.com/wp-content/uploads/2026/03/technijian-logo-full-color-600x125-1.png'

# --- Read Excel ---
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

employees = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
    name = (row[0].value or '').strip().rstrip('-').strip()
    if not name:
        continue
    title = (row[1].value or '').strip()
    email = (row[2].value or '').strip()
    ext = str(row[4].value or '').strip()
    cell_phone = (str(row[6].value or '')).strip()

    # Extract actual hyperlink URL from booking cell (column I)
    booking_cell = row[8]
    booking_url = None
    if booking_cell.hyperlink and booking_cell.hyperlink.target:
        booking_url = booking_cell.hyperlink.target
    else:
        # Fallback: use cell value if it's already a URL
        booking_raw = (str(booking_cell.value or '')).strip()
        if booking_raw.startswith('http'):
            booking_url = booking_raw

    # Determine team
    title_lower = title.lower()
    if 'digital marketing' in title_lower or 'seo' in title_lower:
        team = 'seo'
    elif 'application engineer' in title_lower:
        team = 'dev'
    else:
        team = 'default'

    employees.append({
        'name': name,
        'title': title,
        'email': email,
        'ext': ext,
        'cell': cell_phone,
        'booking_url': booking_url,
        'team': team,
    })


def build_signature(emp):
    """Build a complete HTML email signature for one employee."""

    # Phone line
    office_phone = f'{OFFICE_NUMBER} x{emp["ext"]}' if emp['ext'] else OFFICE_NUMBER
    office_phone_href = re.sub(r'[^0-9]', '', OFFICE_NUMBER) + emp['ext']

    # Cell phone block (optional)
    cell_block = ''
    if emp['cell'] and emp['cell'] != 'None':
        cell_digits = re.sub(r'[^0-9]', '', emp['cell'])
        cell_block = f'''
            <span style="color:#006DB6;font-weight:600;">C:</span> <a href="tel:{cell_digits}" style="color:#59595B;text-decoration:none;">{emp['cell']}</a><br>'''

    # Team booking button
    if emp['team'] == 'seo':
        team_btn_url = SEO_BOOKING_URL
        team_btn_label = 'Book SEO Meeting'
    elif emp['team'] == 'dev':
        team_btn_url = DEV_BOOKING_URL
        team_btn_label = 'Book Dev Meeting'
    else:
        team_btn_url = SUPPORT_BOOKING_URL
        team_btn_label = 'Book time with Support'

    # Personal booking button (optional)
    personal_btn = ''
    if emp['booking_url']:
        personal_btn = f'''              <td style="padding-right:8px;">
                <a href="{emp['booking_url']}"
                  style="display:inline-block;background-color:#006DB6;color:#FFFFFF;font-size:12px;font-weight:600;font-family:'Open Sans','Segoe UI',Helvetica,Arial,sans-serif;padding:6px 16px;border-radius:4px;text-decoration:none;">
                  Book a Meeting
                </a>
              </td>'''

    # Build full HTML
    html = f'''<!-- Technijian Email Signature — {emp['name']} -->
<!-- Template: {emp['team'].upper()} | Generated from employee directory -->

<table cellpadding="0" cellspacing="0" border="0" style="font-family:'Open Sans','Segoe UI',Helvetica,Arial,sans-serif;max-width:600px;">
  <!-- Orange top accent line -->
  <tr>
    <td style="border-top:3px solid #F67D4B;padding-bottom:16px;"></td>
  </tr>
  <tr>
    <!-- Info column -->
    <td style="vertical-align:top;">
      <table cellpadding="0" cellspacing="0" border="0" style="font-family:'Open Sans','Segoe UI',Helvetica,Arial,sans-serif;">
        <!-- Name -->
        <tr>
          <td style="font-size:18px;font-weight:700;color:#1A1A2E;padding-bottom:1px;line-height:1.3;">
            {emp['name']}
          </td>
        </tr>
        <!-- Title -->
        <tr>
          <td style="font-size:13px;color:#006DB6;font-weight:600;padding-bottom:2px;">
            {emp['title']}
          </td>
        </tr>
        <!-- Company -->
        <tr>
          <td style="font-size:12px;color:#59595B;padding-bottom:8px;">
            Technijian
          </td>
        </tr>
        <!-- Contact details -->
        <tr>
          <td style="font-size:12px;color:#59595B;line-height:1.7;">
            <span style="color:#006DB6;font-weight:600;">T:</span> <a href="tel:{office_phone_href}" style="color:#59595B;text-decoration:none;">{office_phone}</a><br>{cell_block}
            <span style="color:#006DB6;font-weight:600;">S:</span> <a href="tel:9493798501" style="color:#59595B;text-decoration:none;">{SUPPORT_NUMBER}</a> <span style="color:#ADB5BD;font-size:11px;">(support)</span><br>
            <span style="color:#006DB6;font-weight:600;">E:</span> <a href="mailto:{emp['email']}" style="color:#006DB6;text-decoration:none;">{emp['email']}</a><br>
            <span style="color:#006DB6;font-weight:600;">W:</span> <a href="https://technijian.com" style="color:#006DB6;text-decoration:none;">technijian.com</a>
          </td>
        </tr>
        <!-- Booking links -->
        <tr>
          <td style="padding-top:8px;padding-bottom:4px;">
            <table cellpadding="0" cellspacing="0" border="0"><tr>
{personal_btn}
              <td>
                <a href="{team_btn_url}"
                  style="display:inline-block;background-color:#F67D4B;color:#FFFFFF;font-size:12px;font-weight:600;font-family:'Open Sans','Segoe UI',Helvetica,Arial,sans-serif;padding:6px 16px;border-radius:4px;text-decoration:none;">
                  {team_btn_label}
                </a>
              </td>
            </tr></table>
          </td>
        </tr>
        <!-- Addresses -->
        <tr>
          <td style="padding-top:6px;font-size:11px;color:#ADB5BD;line-height:1.6;">
            <span style="color:#006DB6;font-weight:600;">USA:</span> 18 Technology Dr., Ste 141, Irvine, CA 92618<br>
            <span style="color:#006DB6;font-weight:600;">India:</span> Plot No. 07, 1st Floor, Panchkula IT Park, Panchkula, Haryana 134109
          </td>
        </tr>
      </table>
    </td>
  </tr>

  <!-- Blue divider line -->
  <tr>
    <td style="padding-top:14px;padding-bottom:10px;">
      <table cellpadding="0" cellspacing="0" border="0" width="100%"><tr><td style="border-top:2px solid #006DB6;"></td></tr></table>
    </td>
  </tr>

  <!-- Logo row -->
  <tr>
    <td style="padding-bottom:10px;">
      <img src="{LOGO_URL}" alt="Technijian - Technology as a Solution" width="160" style="display:block;">
    </td>
  </tr>

  <!-- Social media links (text-based, works without hosted icons) -->
  <tr>
    <td style="padding-bottom:12px;">
      <table cellpadding="0" cellspacing="0" border="0">
        <tr>
          <td style="font-size:11px;font-family:'Open Sans','Segoe UI',Helvetica,Arial,sans-serif;line-height:1;">
            <a href="https://www.linkedin.com/company/technijian" style="color:#006DB6;text-decoration:none;font-weight:600;">LinkedIn</a>
            <span style="color:#E9ECEF;padding:0 6px;">|</span>
            <a href="https://www.facebook.com/Technijian01/" style="color:#006DB6;text-decoration:none;font-weight:600;">Facebook</a>
            <span style="color:#E9ECEF;padding:0 6px;">|</span>
            <a href="https://www.youtube.com/@TechnijianIT" style="color:#006DB6;text-decoration:none;font-weight:600;">YouTube</a>
            <span style="color:#E9ECEF;padding:0 6px;">|</span>
            <a href="https://www.instagram.com/technijianinc/" style="color:#006DB6;text-decoration:none;font-weight:600;">Instagram</a>
            <span style="color:#E9ECEF;padding:0 6px;">|</span>
            <a href="https://twitter.com/technijian_" style="color:#006DB6;text-decoration:none;font-weight:600;">X</a>
            <span style="color:#E9ECEF;padding:0 6px;">|</span>
            <a href="https://www.tiktok.com/@technijian" style="color:#006DB6;text-decoration:none;font-weight:600;">TikTok</a>
            <span style="color:#E9ECEF;padding:0 6px;">|</span>
            <a href="https://in.pinterest.com/technijian01/" style="color:#006DB6;text-decoration:none;font-weight:600;">Pinterest</a>
          </td>
        </tr>
      </table>
    </td>
  </tr>

  <!-- Confidentiality disclaimer -->
  <tr>
    <td style="padding-top:8px;border-top:1px solid #E9ECEF;">
      <p style="font-size:10px;color:#ADB5BD;line-height:1.4;margin:0;font-family:'Open Sans','Segoe UI',Helvetica,Arial,sans-serif;">
        This email and any attachments are confidential and intended solely for the addressee. If you have received this message in error, please notify the sender immediately and delete it from your system. Unauthorized review, use, disclosure, or distribution is prohibited.
      </p>
    </td>
  </tr>
</table>'''
    return html


# --- Generate all signatures ---
os.makedirs(OUTPUT_DIR, exist_ok=True)

for emp in employees:
    html = build_signature(emp)
    # Filename: "Firstname Lastname.html"
    filename = f'{emp["name"]}.html'
    filepath = os.path.join(OUTPUT_DIR, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html)

# Summary
print(f'Generated {len(employees)} signatures in {OUTPUT_DIR}\n')
print(f'{"Name":<30} {"Team":<8} {"Personal Booking":<5}')
print('-' * 50)
for emp in employees:
    has_booking = 'Yes' if emp['booking_url'] else 'No'
    print(f'{emp["name"]:<30} {emp["team"]:<8} {has_booking}')
