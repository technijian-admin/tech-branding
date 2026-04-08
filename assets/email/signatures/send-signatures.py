"""Email each Client Support Engineer their HTML email signature with setup instructions."""
import openpyxl
import win32com.client
import os
import time

# --- Configuration ---
EXCEL_PATH = r'C:\vscode\tech-branding\tech-branding\data\Technijian Directory  (1).xlsx'
SIG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'employees')

# --- Read Excel to get Client Support Engineers ---
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

engineers = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
    name = (row[0].value or '').strip().rstrip('-').strip()
    if not name:
        continue
    title = (row[1].value or '').strip()
    email = (row[2].value or '').strip()

    if 'client support engineer' in title.lower():
        sig_file = os.path.join(SIG_DIR, f'{name}.html')
        if os.path.exists(sig_file):
            engineers.append({'name': name, 'title': title, 'email': email, 'sig_file': sig_file})
        else:
            print(f'WARNING: No signature file for {name} at {sig_file}')

print(f'Found {len(engineers)} Client Support Engineers:\n')
for eng in engineers:
    print(f'  {eng["name"]:<30} {eng["email"]}')

# --- Email body (HTML) ---
EMAIL_SUBJECT = 'Your New Technijian Email Signature'

def build_email_body(name):
    return f"""<html><body style="font-family:'Segoe UI',Arial,sans-serif;font-size:14px;color:#333;">
<p>Hi {name.split()[0]},</p>

<p>Please find your new Technijian email signature attached. Follow the steps below to set it up in Outlook:</p>

<h3 style="color:#006DB6;margin-bottom:4px;">Setup Instructions — Outlook (Desktop)</h3>
<ol>
<li>Open the attached <b>{name}.html</b> file in your web browser (Chrome, Edge, etc.)</li>
<li>Press <b>Ctrl+A</b> to select everything on the page, then <b>Ctrl+C</b> to copy</li>
<li>Open Outlook &rarr; <b>File</b> &rarr; <b>Options</b> &rarr; <b>Mail</b> &rarr; <b>Signatures</b></li>
<li>Click <b>New</b>, name it <b>Technijian</b></li>
<li>In the signature editor, press <b>Ctrl+V</b> to paste</li>
<li>Set it as the default for <b>New messages</b> and <b>Replies/forwards</b></li>
<li>Click <b>OK</b> to save</li>
</ol>

<h3 style="color:#006DB6;margin-bottom:4px;">Setup Instructions — Outlook on the Web (Microsoft 365)</h3>
<ol>
<li>Open the attached <b>{name}.html</b> file in your browser</li>
<li>Press <b>Ctrl+A</b> then <b>Ctrl+C</b> to copy</li>
<li>Go to <a href="https://outlook.office.com">outlook.office.com</a> &rarr; <b>Settings</b> (gear icon) &rarr; <b>View all Outlook settings</b></li>
<li>Navigate to <b>Mail</b> &rarr; <b>Compose and reply</b></li>
<li>Under <b>Email signature</b>, paste your signature (<b>Ctrl+V</b>)</li>
<li>Toggle on <b>"Automatically include my signature on new messages"</b> and <b>"on replies"</b></li>
<li>Click <b>Save</b></li>
</ol>

<p style="color:#59595B;font-size:13px;">If you have any issues, please reach out to IT support.</p>

<p>Thank you,<br>
<b>Technijian IT</b></p>
</body></html>"""


# --- Send emails via Outlook COM ---
print('\nConnecting to Outlook...')
outlook = win32com.client.Dispatch('Outlook.Application')

sent_count = 0
for eng in engineers:
    mail = outlook.CreateItem(0)  # 0 = olMailItem
    mail.To = eng['email']
    mail.Subject = EMAIL_SUBJECT
    mail.HTMLBody = build_email_body(eng['name'])
    mail.Attachments.Add(eng['sig_file'])
    mail.Send()
    sent_count += 1
    print(f'  Sent to {eng["name"]} ({eng["email"]})')
    time.sleep(1)  # Brief pause between sends

print(f'\nDone! Sent {sent_count} emails.')
